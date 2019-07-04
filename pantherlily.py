#Discord
import discord
from discord.ext import commands
from discord import Embed, Game, Guild

#APIs
from utils.apis import d_sync
from utils.apis.discordBotAPI import BotAssist
from utils.apis.ClashConnect import ClashConnectAPI
from utils.apis.rolemanager import Rolemgr
from utils.apis import user_tops

from utils.clash_of_clans import clash_stats
from utils.clash_of_clans import update_donationtable as udt
from utils.help import help_menu
# Database
from utils.database.panther_db import ZuluDB

#New Functions
import aiohttp
import asyncio
import bs4
from collections import OrderedDict
from configparser import ConfigParser
from datetime import datetime, timedelta
from requests import Response
import io
from os import path, listdir
from pathlib import Path
import random
import re
from requests import get
from sys import argv
from sys import exit as ex # Avoid exit built in

# Data visualization
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import json
import coc
import logging

# Set up logging
logger = logging.getLogger('discord')
logger.setLevel(logging.DEBUG)
handler = logging.FileHandler(filename='discord.log', encoding='utf-8', mode='w')
handler.setFormatter(logging.Formatter('%(asctime)s:%(levelname)s:%(name)s: %(message)s'))
logger.addHandler(handler)
#####################################################################################################################
                                             # Set up the environment
#####################################################################################################################
# Look for either the dev or live switch
if len(argv) == 2:
    if argv[-1] == "--live":
        botMode = "liveBot"
    elif argv[-1] == "--dev":
        botMode = "devBot"
    else:
        ex("\n[ERROR] Make sure to add the right switch to activate me.")
else:
    ex("\n[ERROR] Make sure to add the right switch to activate me.")


# Instanciate Config
emoticons = ConfigParser(allow_no_value=True)

if botMode == "liveBot":
    configLoc = 'utils/configurations/panther_conf.json'
    emoticonLoc = 'utils/configurations/emoticons.ini'
    if path.exists(configLoc):
        pass
    else:
        ex(f"Config file does not exist: {configLoc}")
    with open(configLoc) as infile:
        config = json.load(infile)
    emoticons.read(emoticonLoc)
    discord_client = commands.Bot(command_prefix=config[botMode]['bot_prefix'])
    discord_client.remove_command("help")

elif botMode == "devBot":
    configLoc = 'utils/configurations/panther_conf.json'
    emoticonLoc = 'utils/configurations/emoticons.ini'
    if path.exists(configLoc):
        pass
    else:
        ex(f"Config file does not exist: {configLoc}")
    with open(configLoc) as infile:
        config = json.load(infile)
    emoticons.read(emoticonLoc)
    discord_client = commands.Bot(command_prefix=config[botMode]['bot_prefix'])
    discord_client.remove_command("help")


# Instanciate botAssit and DB
dbLoc = config[botMode]['db']
if dbLoc == "None":
    print(f"No dev file set in {botMode}")
    ex("Exiting")
else:
    dbconn = ZuluDB(dbLoc)

botAPI = BotAssist(botMode, configLoc, dbconn, emoticons, config)
coc_client = ClashConnectAPI(config['clash']['zuluclash_token'])
prefx = config[botMode]['bot_prefix'][0]

# instanciate rolemgr
role_mgr = Rolemgr(config)

# coc.py
coc_client2 = coc.Client(config["CoC_API"]["Username"], config["CoC_API"]["Password"])
#####################################################################################################################
                                             # Discord Commands [info]
#####################################################################################################################
@discord_client.event
async def on_ready():
    """
    Simple funciton to display logged in data to terminal
    """
    print(f'\n\nLogged in as: {discord_client.user.name} - {discord_client.user.id}\nDiscord Version: {discord.__version__}\n'
        f"\nRunning in [{botMode}] mode\n"
        "------------------------------------------\n"
        f"Prefix set to:          {prefx}\n"
        f"Config file set to:     {configLoc}\n"
        f"DB File set to:         {dbLoc}\n"
        f"Current exit node:      {get('https://api.ipify.org').text}\n"
        "------------------------------------------")

    game = Game(config[botMode]['game_msg'])
    await discord_client.change_presence(status=discord.Status.online, activity=game)
    guild = discord_client.get_guild(int(config["discord"]["zuludisc_id"]))
    role_mgr.initializer(guild)
    update = d_sync.UpdateLoop(discord_client, dbconn, botMode, coc_client2, config) 
    discord_client.loop.create_task(update.run())

#####################################################################################################################
                                             # Help Menu
#####################################################################################################################
@discord_client.command()
async def help(ctx, option=None):
    # Instanciate the help object
    help_obj = help_menu.Help_Menu(config, botMode)

    # Check conditions to print the help menu
    if option == None:
        await ctx.send(embed=help_obj.utility(True))
    elif option.lower() in ["util", "utility", "u"]:
        await ctx.send(embed=help_obj.utility(True))
    elif option.lower() in ["admin", "administrator", "administration", "u"]:
        await ctx.send(embed=help_obj.administrator(True))
    elif option.lower() in ["accountability", "account", "roster", "acc", "r", "a"]:
        await ctx.send(embed=help_obj.accountability(True))
    elif option.lower() in ["all", "--all", "-all", "-a"]:
        await ctx.send(embed=help_obj.utility(False))
        await ctx.send(embed=help_obj.accountability(False))
        await ctx.send(embed=help_obj.administrator(True))
    else:
        await ctx.send(embed=help_obj.utility(True))

@help.error
async def help_erro(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))
#####################################################################################################################
                                             # Accountability Functions
#####################################################################################################################
@discord_client.command()
async def listroles(ctx):
    """ List the roles and ID in the current channel """
    # Check server and Member Role
    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    tupe = []
    guild_obj = discord_client.get_guild(int(config[botMode]['guild_lock']))
    for i in guild_obj.roles:
        tupe.append((i.name,i.id))

    max_length = 0
    for name in tupe:
        if len(name[0]) > max_length:
            max_length = len(name[0])
    tupe.sort()

    output = ''
    for name in tupe:
        output += "{:<{}} {}\n".format(name[0], max_length, name[1])

    await ctx.send("```{}```".format(output))

@listroles.error
async def listroles_error(ctx, error):
    await ctx.send(embed=discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))

@discord_client.command()
async def lcm(ctx):
    """ List users in the current clan with their tag """

    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    res = coc_client.get_clan(config['clash']['zulu'])
    # get all mems from DB
    db_result = dbconn.get_all_active()

    strength = {
        12 : 0,
        11 : 0,
        10 : 0,
         9 : 0,
         8 : 0
    }
    for row in db_result:
        strength[int(row[2])] = strength[int(row[2])] + 1

    # Quick check to  make sure that the https request was good
    if int(res.status_code) != 200:
        embed = Embed(color=0xff0000)
        msg = (f"Bad HTTPS request, please make sure that the bots IP is in the CoC whitelist. "
        f"Our current exit node is {get('https://api.ipify.org').text}")
        embed.add_field(name="Bad Request: {}".format(res.status_code),value=msg)
        await ctx.send(embed=embed)
        return

    # If we're able to talk to COC api
    else:
        mem_list = []
        for user in res.json()['memberList']:
            mem_list.append((user['name'],user['tag']))

        #sort the list and get the max length
        max_length = 0
        for user in mem_list:
            if len(user[0]) > max_length:
                max_length = len(user[0])
        mem_list.sort(key = lambda tupe_item: tupe_item[0].lower())
        output = ''
        for index, user in enumerate(mem_list):
            output += "[{:>2}] {:<{}} {}\n".format(index+1, user[0], max_length, user[1])

        view = await ctx.send("Current Members in Reddit Zulu\n```{}```".format(output))
        await view.add_reaction(emoticons["tracker bot"]["plus"].lstrip("<").rstrip(">"))


        # Reaction 
        # Check function -- uses view variable that is scoped later   
        def check(reaction, user):
            # Make sure that the reaction is for the correct message 
            if view.id == reaction.message.id:
                return user.bot == False
            else:
                return False

        try:
            await ctx.bot.wait_for('reaction_add', timeout = 60, check=check)
            await view.clear_reactions()
            embed = discord.Embed(title="Reddit Zulu Strength", color=0x00FF80)
            embed.add_field(name="Registered Members: ", value=len(db_result), inline=False)
            embed.add_field(name="TH12s: ", value=strength[12], inline=False)
            embed.add_field(name="TH11s: ", value=strength[11], inline=False)
            embed.add_field(name="TH10s: ", value=strength[10], inline=False)
            embed.add_field(name="TH9s: ", value=strength[9], inline=False)
            await ctx.send(embed=embed)
            return
        except asyncio.TimeoutError:
            await view.clear_reactions()
            return


@lcm.error
async def lcm_error(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))

@discord_client.command()
async def roster(ctx):
    """ Function is used to check what members are in which server """
    if await botAPI.rightServer(ctx, config):
        pass
    else:
        return

    # get all clan members
    res = coc_client.get_clan(config['clash']['zulu'])

    # Quick check to  make sure that the https request was good
    if int(res.status_code) != 200:
        embed = Embed(color=0xff0000)
        msg = (f"Bad HTTPS request, please make sure that the bots IP is in the CoC whitelist. "
        f"Our current exit node is {get('https://api.ipify.org').text}")
        embed.add_field(name="Bad Request: {}".format(res.status_code),value=msg)
        await ctx.send(embed=embed)
        return

    zuluServer = discord_client.get_guild(int(config['discord']['zuludisc_id']))
    zbpServer = discord_client.get_guild(int(config['discord']['plandisc_id']))

    if zuluServer == None or zbpServer == None:
        await ctx.send("Unable to instantiate the guild object")
        return

    roster = {}
    # for zMember in (mem for mem in zuluServer.members if 'CoC Members' in (role.name for role in mem.roles)):
    mems = [ mem for mem in zuluServer.members if 'CoC Members' in (role.name for role in mem.roles) and mem.name != "ZuluTest" ]
    mems.sort(key=lambda x: x.display_name.lower())
    for zMember in mems:
        roster[zMember.display_name] = {
            "Clash"       :   False,
            "zuluServer"  :   True,
            "zbpServer"   :   False,
            "database"    :   False
        }
         # check if member is in zbpServer
        if zMember.id in ( pMember.id for pMember in zbpServer.members ):
            roster[zMember.display_name]['zbpServer'] = True

        queryResult = dbconn.get_user_byDiscID((zMember.id,))
        if len(queryResult) == 1:
            roster[zMember.display_name]['database'] = True

            if queryResult[0][0] in ( member['tag'] for member in res.json()['memberList'] ):
                roster[zMember.display_name]['Clash'] = True


    line = (f"{emoticons['tracker bot']['zuluServer']}{emoticons['tracker bot']['planningServer']}{emoticons['tracker bot']['redditzulu']}{emoticons['tracker bot']['database']}\u0080\n")
    mem_count = 1
    for userName in roster.keys():
        if roster[userName]['zuluServer'] == True:
            line += f"{emoticons['tracker bot']['true']}"
        else:
            line += f"{emoticons['tracker bot']['false']}"

        if roster[userName]['zbpServer'] == True:
            line += f"{emoticons['tracker bot']['true']}"
        else:
            line += f"{emoticons['tracker bot']['false']}"

        if roster[userName]['Clash'] == True:
            line += f"{emoticons['tracker bot']['true']}"
        else:
            line += f"{emoticons['tracker bot']['false']}"

        if roster[userName]['database'] == True:
            line += f"{emoticons['tracker bot']['true']}"
        else:
            line += f"{emoticons['tracker bot']['false']}"

        line += f"  **{mem_count:>2}**  {userName}\n"
        mem_count += 1
        if len(line) > 1700:
            await ctx.send(line)
            line = ''
    if line != '':
        await ctx.send(line)
    legend = (f"{emoticons['tracker bot']['zuluServer']} Member is in Reddit Zulu discord.\n"
            f"{emoticons['tracker bot']['planningServer']} Member is in ZBP discord.\n"
            f"{emoticons['tracker bot']['redditzulu']} Member is currently in Reddit Zulu in-game.\n"
            f"{emoticons['tracker bot']['database']} Member is registered with PantherLily.\n"
            f"{emoticons['tracker bot']['plus']} Locate all users in real-time, may take 15 seconds.")
    view = await ctx.send(f"**LEGEND**\n{legend}\n**WARNING**\nClash query is not performed if user is missing from the database. Use `{prefx}lcm` "
        "to get an up to date list of clan members.")

    # Add reaction button
    await view.add_reaction(emoticons["tracker bot"]["plus"].lstrip("<").rstrip(">"))

    # Check for click
    def check(reaction, user):
        # Make sure that the reaction is for the correct message 
        if view.id == reaction.message.id:
            return user.bot == False
        else:
            return False

    try:
        await ctx.bot.wait_for('reaction_add', timeout = 10, check=check)
        await ctx.send("**[!] **Please hold while I get that for ya...")
        await view.clear_reactions()
        async with ctx.typing():
            # Dictionary of clans
            user_distribution = {
                "#P0Q8VRC8" : [],
                "#2Y28CGP8" : [],
                "#8YGOCQRY" : [],
                "Unknown" : []
            }
            # get all active users from database 
            active_members = [ tag[0] for tag in dbconn.get_all_active() ]
            async for player in coc_client2.get_players(active_members):
                try:
                    user_distribution[player.clan.tag].append((
                        player.name,
                        player.town_hall
                    ))
                except:
                    if player.clan:
                        clan = player.clan.name
                    else:
                        clan = "No Clan"
                    user_distribution["Unknown"].append((
                        player.name,
                        clan
                    ))
            
            # Sort the list
            for section in user_distribution.keys():
                th12, th11, th10, th9 = [], [], [], []
                user_distribution[section].sort(key = lambda x: x[0].lower())
                for user in user_distribution[section]:
                    if user[1] == 12:
                        th12.append(user)
                    elif user[1] == 11:
                        th11.append(user)
                    elif user[1] == 10:
                        th10.append(user)
                    else: 
                        th9.append(user)
                order = []
                order.extend(th12)
                order.extend(th11)
                order.extend(th10)
                order.extend(th9)
                user_distribution[section] = order


            # Create the outputs
            if user_distribution[f"#{config['clash']['misfits']}"]:
                misfits_out = "**Reddit Misfits:**\n"
                for member in user_distribution[f"#{config['clash']['misfits']}"]:
                    misfits_out += f"`{member[1]:<4}{member[0]}`\n"
                misfits_out += f"""Count: {len(user_distribution[f"#{config['clash']['misfits']}"])}/{len(active_members)}"""
                await ctx.send(misfits_out)

            if user_distribution[f"#{config['clash']['elephino']}"]:
                elephino_out = "**Reddit Elephino:**\n"
                for member in user_distribution[f"#{config['clash']['elephino']}"]:
                    elephino_out += f"`{member[1]:<4}{member[0]}`\n"
                elephino_out += f"""Count: {len(user_distribution[f"#{config['clash']['elephino']}"])}/{len(active_members)}"""
                await ctx.send(elephino_out)

            if user_distribution[f"#{config['clash']['zulu']}"]:
                zulu_out = "**Reddit Zulu:**\n"
                for member in user_distribution[f"#{config['clash']['zulu']}"]:
                    zulu_out += f"`{member[1]:<4}{member[0]}`\n"
                zulu_out += f"""Count: {len(user_distribution[f"#{config['clash']['zulu']}"])}/{len(active_members)}"""
                await ctx.send(zulu_out)

            if user_distribution["Unknown"]:
                unknown_out = "**Users not in our clans**:\n"
                for member in user_distribution["Unknown"]:
                    unknown_out += f"`{member[0]:<24}{member[1]}`\n"
                unknown_out += f"""Count: {len(user_distribution["Unknown"])}/{len(active_members)}"""
                await ctx.send(unknown_out)

    except asyncio.TimeoutError:
        await view.clear_reactions()
@roster.error
async def roster_error(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))


#####################################################################################################################
                                             # Commands for all users
#####################################################################################################################
@discord_client.command()
async def invite(ctx, *arg):
    """ Get the channel object to use the invite method of that channel """

    if await botAPI.rightServer(ctx, config):
        targetServer = int(config['discord']['plandisc_id'])
        targetChannel = int(config['discord']['plandisc_channel'])

    else:
        return

    # Try to create the invite object
    if len(arg) == 1 and arg[0].isdigit():
        channel = botAPI.invite(discord_client, targetServer, targetChannel)
        inv = await channel.create_invite(max_age = (int(arg[0]) *60), max_uses = 1 )
        await ctx.send(inv)
        return

    elif len(arg) == 0:
        channel = botAPI.invite(discord_client, targetServer, targetChannel)
        inv = await channel.create_invite(max_age = 600, max_uses = 1 )
        await ctx.send(inv)
        return

    else:
        await ctx.send("Wrong arguments used")
        return

@invite.error
async def newinvite_error(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))

@discord_client.command(aliases=["man"])
async def manual(ctx):
    f = discord.File("utils/configurations/PantherLily_V1.pdf", filename="Manual.pdf")
    await ctx.send(file=f)

@discord_client.command(aliases=["s"])
async def stats(ctx, *, user=None):
    _max = False
    if user:
        for opt in ["--max", "-m"]:
            if user.split(" ")[-1].lower() == opt:
                if len(user.split(" ")) == 1:
                    user = None
                    _max = True
                    break
                else:
                    user = user.split(" "+opt)[0]
                    _max = True
                    break
    if user == None:
        user = ctx.author
        userID = user.id
        result = dbconn.get_user_byDiscID((userID,))
        if len(result) == 0:
            await ctx.send(f"No data was found for: {ctx.author.display_name}")
            return
    else:
        userID = await botAPI.user_converter_db(ctx, user)
        if userID == None:
            await ctx.send(f"No data was found for **{user}**")
            return
        result = dbconn.get_user_byDiscID((userID,))
        if len(result) == 0:
            await ctx.send(f"No data was found for {user.display_name}")
            return

    if len(result) == 0:
        msg = (f"Could not find {ctx.author.display_name} in Zulu's database. Make sure they have "
        "been added.")
        await ctx.send(embed = Embed(title=f"SQL ERROR", description=msg, color=0xff0000))
        return

    elif len(result) > 1:
        msg = (f"Found duplicate discord ID entries")
        await ctx.send(embed = Embed(title=f"SQL ERROR", description=msg, color=0xff0000))
        return

    player = await coc_client2.get_player(result[0][0], cache=False)

    if player == None:
        msg = (f"Bad HTTPS request, please make sure that the bots IP is in the CoC whitelist. "
        f"Our current exit node is {get('https://api.ipify.org').text}")
        await ctx.send(embed = Embed(title=f"HTTP", description=msg, color=0xff0000))
        return

    # Get display objects
    desc, troopLevels, spellLevels, heroLevels, gains, sieges = clash_stats.stat_stitcher(player, emoticonLoc, _max)
    embed = Embed(title = f"**__{player.name}__**", description=desc, color = 0x000080)
    embed.add_field(name="**Gains**", value=gains, inline = False)
    embed.add_field(name = "**Heroes**", value=heroLevels, inline = False)
    embed.add_field(name = "**Troops**", value=troopLevels, inline = False)
    embed.add_field(name = "**Spells**", value=spellLevels, inline = False)
    if sieges != None:
        embed.add_field(name = "**Sieges**", value=sieges, inline = False)
    embed.set_footer(text=config[botMode]["version"]+" "+config[botMode]["panther_url"])
    if player.league.badge.small == None:
        f = discord.File("Images/Unranked_League.png", filename='unrank.png')
        embed.set_thumbnail(url="attachment://unrank.png")
        await ctx.send(embed=embed, file=f)
    else:
        embed.set_thumbnail(url=player.league.badge.small)
        await ctx.send(embed=embed)

@stats.error
async def stats_error(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))

@discord_client.command(aliases=["d"])
async def donation(ctx, *, user=None):
    """
    Discord command is used to display a users donation status. It queries the database for their current donation and 
    performs a comparison between the start of the week and current day and supplies the difference.
    
    Arguments:
        ctx {Discord Context} -- Message context from a discord chat
        user {str or discord obj} -- User to query against (Default: None)

    Returns:
        Discord message using ctx.message
    """  
    if user == None:
        query_result = dbconn.get_user_byDiscID((ctx.author.id,))
        if len(query_result) == 0:
            await ctx.send(f"No data was found for {ctx.author.display_name}")
            return
    else:
        discord_member = await botAPI.user_converter_db(ctx, user)
        if discord_member == None:
            msg = f"User: {user} not found."
            await botAPI.await_error(ctx, msg, "USER NOT FOUND")
            return
    
        else:
            if discord_member != None:
                query_result = dbconn.get_user_byDiscID((discord_member,))
                if len(query_result) == 0:
                    await ctx.send(f"No data was found for {user}")
                    return
                else:
                    pass
            else:
                msg = f"Was not able to find a user with the argument of: {user}"
                await botAPI.await_error(ctx, msg, "USER NOT FOUND")
                return


    if not query_result:
        msg = (f"{ctx.author.display_name} was not found in our database. Have they been added?")
        await ctx.send(embed = discord.Embed(title="SQL ERROR", description=msg, color=0xFF0000))
        return

    elif len(query_result) > 1:
        users = [ i[1] for i in query_result ]
        msg = (f"Oh oh, looks like we have duplicate entries with the same discord ID. Users list: {users}")
        await ctx.send(embed = discord.Embed(title="SQL ERROR", description=msg, color=0xFF0000))
        return

    elif query_result[0][7] == "False":
        msg = (f"Sorry {query_result[0][1]}, I am no longer tracking your donations as your enrollment to Reddit Zulu is set to False. "
        "Please ping @CoC Leadership if this is a mistake.")
        await ctx.send(embed = discord.Embed(title="SQL ERROR", description=msg, color=0xFF0000))
        return

    # Update users donation
    await udt.update_user(dbconn, coc_client2, query_result[0][0])

    lastSun = botAPI.last_sunday()
    nextSun = lastSun + timedelta(days=7)
    donation = dbconn.get_Donations((query_result[0][0], lastSun.strftime("%Y-%m-%d %H:%M:%S"), nextSun.strftime("%Y-%m-%d %H:%M:%S")))
    try:
        lastDon = datetime.strptime(donation[0][0], "%Y-%m-%d %H:%M:%S")
    except:
        await ctx.send("Please wait an hour to accurately calculate your donations. Thank you for your patience.")
        return


    if len(donation) > 2:
        val = (lastDon - lastSun)
        if val.days == 0:
            remain = nextSun - datetime.utcnow()
            day = remain.days
            time = str(timedelta(seconds=remain.seconds)).split(":")
            msg = (f"**Donation Stat:**\n{donation[-1][2] - donation[0][2]} | 300\n"
                f"**Time Remaining:**\n{day} days {time[0]} hours {time[1]} minutes")
            embed = discord.Embed(title=f"__**{query_result[0][1]}**__", description=msg, color=0x000080)
            embed.set_footer(text=config[botMode]["version"])
            await ctx.send(embed=embed)
            return

        else:
            active = datetime.utcnow() - lastDon
            await ctx.send(f"**WARNING**\nOnly {active.days} day(s) of data have been recorded. \nFirst donation on: [{lastDon.strftime('%Y-%m-%d %H:%M:%S')} Zulu]")

            remain = nextSun - datetime.utcnow()
            day = remain.days
            time = str(timedelta(seconds=remain.seconds)).split(":")
            msg = (f"**Donation Stat:**\n{donation[-1][2] - donation[0][2]} | 300\n"
                f"**Time Remaining:**\n{day} days {time[0]} hours {time[1]} minutes")
            embed = discord.Embed(title=f"__**{query_result[0][1]}**__", description=msg, color=0x000080)
            embed.set_footer(text=config[botMode]["version"])
            await ctx.send(embed=embed)
            return

    else:
        await ctx.send("No data was returned, try running me again.")

@donation.error
async def mydonations_error(ctx, error):
    await botAPI.await_error(ctx, error.__str__(), "RUNTIME ERROR")

#####################################################################################################################
                                             # Admin Commands
#####################################################################################################################
@discord_client.command(aliases=["add_user"])
async def user_add(ctx, clash_tag, *, disc_mention, fin_override=None):
    """
    Function to add a user to the database and initiate tracking of that user
    """
    # Check server and Member Role
    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    # try to get the user object 
    disc_user_id = await botAPI.user_converter_db(ctx, disc_mention)
    disc_user_obj = await botAPI.userConverter(ctx, disc_user_id)

    # Break if the user does not exist 
    if disc_user_obj == None:
        msg = (f"User id {disc_mention} does not exist on this server.")
        await ctx.send(embed = Embed(title="ERROR", description=msg, color=0xFF0000))
        return

    # Query CoC API to see if we have the right token and the right tag
    try:
        player = await coc_client2.get_player(f"#{clash_tag.lstrip('#')}")
    except coc.errors.NotFound as exception:
        player = None
    
    # Handle HTTP error
    if player == None:
        msg = (f"Was not able to find {clash_tag} in CoC servers.") 
        await ctx.send(embed = Embed(title="HTTP ERROR", description=msg, color=0xFF0000))
        return

    # Retrieve the roles
    role_cocmember = role_mgr.get_role("CoC Members")
    role_thlvl = role_mgr.get_th(player.town_hall)

    if role_cocmember == None:
        msg = (f"Clash role [CoC Members] was not found in Reddit Zulu discord")
        await botAPI.await_error(msg, "INSTANCE ERROR")
        return
    if role_thlvl == None:
        msg = (f"Clash role [th{player.town_hall}s] was not found in Reddit Zulu discord")
        await botAPI.await_error(msg, "INSTANCE ERROR")
        return

    # Add roles to a role list
    role_list = [role_cocmember, role_thlvl]

    # Change users default roles
    msg = (f"Applying default roles to {player.name}")
    await ctx.send(embed = Embed(description=msg, color=0x5c0189))
    try:
        await disc_user_obj.add_roles(*role_list)
        await ctx.send(f"[+] {role_cocmember.name}")
        await ctx.send(f"[+] {role_thlvl.name}")
        await ctx.send(embed = Embed(description="Roles applied", color=0x00ff00))
    except:
        await botAPI.await_error(ctx, "Could not add roles")
        

    msg = (f"Changing {player.name}'s nickname to reflect their in-game name.")
    await ctx.send(embed = Embed(title=msg, color=0x5c0189))

    # Change users nickname
    if disc_user_obj.display_name == player.name:
        msg = (f"{player.name}'s discord nickname already reflects their in-game name.")
        await ctx.send(embed = Embed(description=msg, color=0xFFFF00))
    else:
        oldName = disc_user_obj.display_name
        try:
            await role_mgr.change_name(disc_user_obj, player.name)
            msg = (f"Changed {player.name} discord nickname from {oldName} to {disc_user_obj.display_name}")
            await ctx.send(embed = Embed(description=msg, color=0x00ff00))
        except:
            msg = (f"It is impossible for a mere bot to change the nickname of a boss like you. "
            "Seriously though, bots are prohibited from doing this action to a discord leader.")
            await ctx.send(embed = Embed(description=msg, color=0xff0000))


    # Add user to database
    msg = (f"Adding {player.name} to Reddit Zulu's database.")
    await ctx.send(embed = Embed(title=msg, color=0x5c0189))
    error = dbconn.insert_userdata((
        player.tag,
        player.name,
        player.town_hall,
        player.league.name,
        disc_user_obj.id,
        disc_user_obj.joined_at.strftime('%Y-%m-%d %H:%M:%S'),
        "False",
        "True",
        "",
    ))
    if error != None:
        if error.args[0] == "UNIQUE constraint failed: MembersTable.Tag":
            msg = (f"UNIQUE constraint failed: MembersTable.Tag: {player.tag}\n\nUser already exists. Attempting to re-activate {player.name}")
            await ctx.send(embed = Embed(description=msg, color=0xFFFF00))
            result = dbconn.is_Active((player.tag))
            if isinstance(result, str):
                await ctx.send(embed = Embed(title="SQL ERROR", description=result, color=0xFF0000))
                return

            elif result[7] == "True": # If activ
                msg = (f"{player.name} is already set to active in the database.")
                await ctx.send(embed = Embed(title="SQL ERROR", description=msg, color=0xFF0000))
                return
            else:
                result = dbconn.set_Active(("True", player.tag))

                if isinstance(result, str):
                    await ctx.send(embed = Embed(title="SQL ERROR", description=result, color=0xFF0000))
                    return
                else:
                    msg = (f"Successfully set {player.name} to active")
                    await ctx.send(embed = Embed(description=msg, color=0x00FF00))
        else:
            await ctx.send(embed = Embed(title="SQL ERROR", description=error.args[0], color=0xFF0000)) #send.args[0] == "database is locked":
            return


    msg = (f"{disc_user_obj.display_name} added. Please copy and paste the following output into #sidekick-war-caller")
    await ctx.send(embed = Embed(description=msg, color=0x00FF00))
    await ctx.send(f"/add {clash_tag.upper()} {disc_user_obj.mention}")
    return

@user_add.error
async def info_error(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))

@discord_client.command(aliases=["remove_user", "user_disable", "disable_user"])
async def user_remove(ctx, *, query, suppress=None, note_to_add=None):

    # Check server and Member Role
    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    if "-m" in query:
        arguments = query.split(" -m ")
        query = arguments[0]
        suppress = True
        note_to_add = arguments[-1]

    # Attempt to resolve the user name
    member_id = await botAPI.user_converter_db(ctx, query)
    if member_id == None:
        desc = (f"Was unable to resolve {query}. This command supports mentions, "
            "IDs, username and nicknames.")
        await ctx.send(embed = discord.Embed(title="RESOLVE ERROR", description=desc, color=0xFF0000))
        return
    
    # try to get the users member object 
    user_obj = ctx.guild.get_member(member_id)

    # See if you can get the member object
    if isinstance(user_obj, discord.Member):
        user_name = user_obj.display_name
    else:
        user_name = query

    if suppress==None and note_to_add==None:
        msg = (f"You are about to set {user_name} CoC Member status to False "
            "are you sure you would like to continue?\n(Yes/No)")
        await ctx.send(msg)

        response = await discord_client.wait_for('message', check = botAPI.yesno_check)
        if response.content.lower() == 'no':
            await ctx.send("Terminating function")
            return

        example = (f"SgtMajorDoobie failed to meet weekly donation quota\n\n"
            "msgID:546408720872112128\nmsgID:546408729155993615")

        await ctx.send("A kick message is required. You are able to enter any text you "
            f"like or message IDs. You can then use {prefx}retrieve_msg command to extract "
            "any message IDs you have included in this kick message. To include a message "
            "ID make sure to prefix the ID with msgID:<id> to make it easier to parse for you.\n\n**Example:**")
            
        await ctx.send("```\n"
            f"{example}\n"
            "```")

        await ctx.send("Please enter your message:")

        def check(m):
            return m.author.id == ctx.author.id

        response = await discord_client.wait_for('message', check=check)
        await ctx.send(f"**You have entered:**\n{response.content}\n\nContinue? (Yes/No)")

        response2 = await discord_client.wait_for('message', check = botAPI.yesno_check)
        if response2.content.lower() == 'no':
            await ctx.send("Terminating function")
            return

        # Re query just incase we got the user from the Discord API
        oldNote = dbconn.get_user_byDiscID((member_id,))

        note = oldNote[0][8]
        note += f"\n\n[{datetime.utcnow().strftime('%d-%b-%Y %H:%M').upper()}]\nNote by {ctx.author.display_name}\n"
        note += f"{response.content}"
        result = dbconn.set_kickNote((note, "False", member_id,))
        if result == 1:
            desc = (f"Successfully set {oldNote[0][1]} active status to False with "
                "the note provided above.")
            await ctx.send(embed = discord.Embed(title="COMMIT SUCCESS", description=desc, color=0x00FF00))  
        
        else:
            desc = (f"Unable to find {oldNote[1]} in the database. Use {prefx}roster to verify "
                "user.")
            await ctx.send(embed = discord.Embed(title="SQL ERROR", description=desc, color=0xFF0000))

    elif suppress and note_to_add:      
        # Re query just incase we got the user from the Discord API
        oldNote = dbconn.get_user_byDiscID((member_id,))

        note = oldNote[0][8]
        note += f"\n\n[{datetime.utcnow().strftime('%d-%b-%Y %H:%M').upper()}]\nNote by {ctx.author.display_name}\n"
        note += f"{note_to_add}"
        result = dbconn.set_kickNote((note, "False", member_id,))
        if result == 1:
            desc = (f"Successfully set {oldNote[0][1]} active status to False with "
                "the note provided above.")
            await ctx.send(embed = discord.Embed(title="COMMIT SUCCESS", description=desc, color=0x00FF00))  
        
        else:
            desc = (f"Unable to find {oldNote[1]} in the database. Use {prefx}roster to verify "
                "user.")
            await ctx.send(embed = discord.Embed(title="SQL ERROR", description=desc, color=0xFF0000))
    
    # Remove user roles
    await ctx.send("Attempting to remove roles...") 
    try:
        user = ctx.guild.get_member(member_id)
        remove_roles = [
            303965219829448705,
            455572149277687809,
            303965664375472128,
            303965505813872641,
            540615053855162408,
            540615255013851136,
            294287799010590720,
            297113442618179585
        ]
        role_objects = [ ctx.guild.get_role(role) for role in remove_roles ]
        await user.remove_roles(*role_objects)
        await ctx.send("Removed roles")
    except: 
        await ctx.send("Could not remove roles from the user")        
    

@user_remove.error
async def kickuser_error(ctx, error):
    await botAPI.await_error(ctx, error.__str__(),"RUNTIME ERROR")

@discord_client.command()
async def addnote(ctx, *, mem):
    # User and Server check
    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    # Get user object 
    member_id = await botAPI.user_converter_db(ctx, mem)

    # If user object can't re resolved then exit 
    if member_id == None:
        desc = (f"Was unable to resolve {mem}. This command supports mentions, "
        "IDs, username and nicknames.")
        await ctx.send(embed = discord.Embed(title="RESOLVE ERROR", description=desc, color=0xFF0000))
        return

    # Query the users data 
    discord_member = discord_client.get_user(member_id)
    if discord_member != None:
        display_name = discord_member.display_name
    else:
        display_name = "user"

    example = (f"Missed attack\nmsgID:123456789654\nmsgID: 4654876135")
    await ctx.send(f"What would you like to add {ctx.author.display_name}? "
        f"Remember to use the 'msgID:' when you want to include message ids in your notes.\n**Example**\n")
            
    await ctx.send("```\n"
        f"{example}\n"
        "```")

    def check(m):
        return m.author.id == ctx.author.id

    response = await discord_client.wait_for('message', check=check)
    await ctx.send(f"**You have entered:**\n{response.content}\n\nContinue? (Yes/No)")

    response2 = await discord_client.wait_for('message', check = botAPI.yesno_check)
    if response2.content.lower() == 'no':
        await ctx.send("Terminating function")
        return

    oldNote = dbconn.get_user_byDiscID((member_id,))
    note = oldNote[0][8]
    note += f"\n\n[{datetime.utcnow().strftime('%d-%b-%Y %H:%M').upper()}]\nNote by {ctx.author.display_name}\n"
    note += f"{response.content}"
    result = dbconn.set_kickNote((note, "True", member_id,))
    if result == 1:
        desc = (f"Successfully added a note to {display_name}")
        await ctx.send(embed = discord.Embed(title="COMMIT SUCCESS", description=desc, color=0x00FF00))
        return  
    
    else:
        desc = (f"Unable to find {display_name} in the database. Use {prefx}roster to verify "
            "user.")
        await ctx.send(embed = discord.Embed(title="SQL ERROR", description=desc, color=0xFF0000))
        return 

@discord_client.command()
async def lookup(ctx, option, *, query):
    # Check server and Member Role
    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    if option not in ['--tag', '-t', '--name', '-n', '--global', '-g']:
        desc = (f"Invalid argument supplied: {option}")
        await ctx.send(embed = discord.Embed(title="ARG ERROR", description=desc, color=0xFF0000))
        return
    
    # disable showing notes if user is not in coc_head chats
    inLeaderChat = False
    if ctx.channel.id in [293953660059385857, 293953660059385857, 498720245691973672, 331565220688297995, 503660106110730256]: #513334681354240000  
        inLeaderChat = True

    # Check function -- uses view variable that is scoped later   
    def check(reaction, user):
        # Make sure that the reaction is for the correct message 
        if view.id == reaction.message.id:
            return user.bot == False
        else:
            return False

    # -tag option
    if option in ['--tag', '-t']:
        # Add a hash tag for the user if it's not there
        if query.startswith("#"):
            tag = query
        else:
            tag = f"#{query}"

        # Query the tab looking for this tag
        results = dbconn.get_user_byTag((tag,))
        if len(results) > 0:
            for result in results:
                if result[8] == '':
                    note = "Empty"
                else:
                    note = result[8]
                embed = discord.Embed(title=result[1], color=0x00FF80)
                embed.add_field(name="ClashTag:", value=result[0], inline=False)
                embed.add_field(name="TownHallLevel:", value=result[2], inline=False)
                embed.add_field(name="DiscordID:", value=result[4], inline=False)
                embed.add_field(name="Database Join:", value=result[5], inline=False)
                if result[7] == "True":
                    active = "Active"
                else:
                    active = "Inactive"
                embed.add_field(name="Status:", value=active, inline=False)
                if inLeaderChat:
                    embed.add_field(name="Profile Note:", value=note, inline=False)
                else:
                    embed.add_field(name="Profile Note:", value="Disabled in this channel.", inline=False)
                view = await ctx.send(embed=embed)
                await view.add_reaction(emoticons["tracker bot"]["plus"].lstrip("<").rstrip(">"))

                try:
                    await ctx.bot.wait_for('reaction_add', timeout = 60, check=check)
                    await view.clear_reactions()
                    await ctx.send(result[4])
                    await ctx.send(result[0])
                except asyncio.TimeoutError:
                    await view.clear_reactions()
            return
        else:
            desc = (f"No results found in the database using ClashTag: {tag}")
            await ctx.send(embed = discord.Embed(title="RECORD NOT FOUND", description=desc, color=0xFF0000))
            return

    if option in ['--name', '-n']:
        # Attempt to resolve the user name
        member_id = await botAPI.user_converter_db(ctx, query)

        # If can't find the user
        if member_id == None:
            await ctx.send("Could not find that user in the database.")
            return

        # Query the users table
        result = dbconn.get_user_byDiscID((member_id,))
        if len(result) > 0:
            result = result[0]
            if result[8] == '':
                note = "Empty"
            else:
                note = result[8]
            embed = discord.Embed(title=result[1], color=0x00FF80)
            embed.add_field(name="ClashTag:", value=result[0], inline=False)
            embed.add_field(name="TownHallLevel:", value=result[2], inline=False)
            embed.add_field(name="DiscordID:", value=result[4], inline=False)
            embed.add_field(name="Database Join:", value=result[5], inline=False)
            if result[7] == "True":
                active = "Active"
            else:
                active = "Inactive"
            embed.add_field(name="Status:", value=active, inline=False)
            if inLeaderChat:
                embed.add_field(name="Profile Note:", value=note, inline=False)
            else:
                embed.add_field(name="Profile Note:", value="Disabled in this channel.", inline=False)

            # Reaction  
            view = await ctx.send(embed=embed)
            await view.add_reaction(emoticons["tracker bot"]["plus"].lstrip("<").rstrip(">"))
            try:
                await ctx.bot.wait_for('reaction_add', timeout = 60, check=check)
                await view.clear_reactions()
                await ctx.send(result[4])
                await ctx.send(result[0])
            except asyncio.TimeoutError:
                await view.clear_reactions()
        return

    if option in ['--global', '-g']:
        res = await botAPI.userConverter(ctx, query)
        if res == None:
            await ctx.send("Could not find user in this server")
        else: 
            embed = discord.Embed(title="Global Lookup", color=0x00FF80)
            embed.add_field(name="Username: ", value=res, inline=False)
            embed.add_field(name="Display: ", value=res.display_name, inline=False)
            embed.add_field(name="Discord ID: ", value=res.id, inline=False)
            embed.add_field(name="Joined Server: ", value=res.joined_at.strftime("%d %b %Y %H:%M:%S").upper(), inline=False)
            out = ''
            for i in res.roles:
                out += f"{i.name}\n"
            embed.add_field(name="Current Roles: ", value=out, inline=False)

            # Reactions
            view = await ctx.send(embed=embed)
            await view.add_reaction(emoticons["tracker bot"]["plus"].lstrip("<").rstrip(">"))
            try:
                await ctx.bot.wait_for('reaction_add', timeout = 60, check=check)
                await view.clear_reactions()
                await ctx.send(res.id)
                return
            except asyncio.TimeoutError:
                await view.clear_reactions()
                return

@lookup.error
async def search_error(ctx, error):
    embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000)
    embed.add_field(name=f"**{prefx}lookup** <--__name__ | --__tag__ | --__discordID__> <__argument__>", value="See help menu")  
    await ctx.send(embed=embed)


@discord_client.command()
async def deletenote(ctx, *, mem):
    """ Function used to delete notes for the user supplieds database """

    # Check server and User
    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    # Get user object 
    member_id = await botAPI.user_converter_db(ctx, mem)

    # If user object can't re resolved then exit 
    if member_id == None:
        desc = (f"Was unable to resolve {mem}. This command supports mentions, "
        "IDs, username and nicknames.")
        await ctx.send(embed = discord.Embed(title="RESOLVE ERROR", description=desc, color=0xFF0000))
        return

    result = dbconn.get_user_byDiscID((member_id,))
    if len(result) == 1:
        user_obj = await ctx.guild.get_member(member_id)

        # See if you can get the member object
        if isinstance(user_obj, discord.Member):
            user_name = user_obj.display_name
        else:
            user_name = mem

        note = result[0][8]
        await ctx.send(f"The current note set for {user_name} is:\n\n```{note}```\n\n ")

        await ctx.send("```\n"
            f"{note}\n"
            "```")

        await ctx.send("Would you like to proceed with deleting this note? This action cannot be undone.\n(Yes/No)")

        response = await discord_client.wait_for('message', check = botAPI.yesno_check)
        if response.content.lower() == "no":
            await ctx.send("Terminating function")
            return
        else: 
            note = f"[{datetime.utcnow().strftime('%d-%b-%Y %H:%M').upper()}]\nDeleted by {ctx.author.display_name}"
            res = dbconn.set_kickNote((note, result[0][7], member_id))
            if res == 1:
                desc = (f"Successfully cleared {user_name} note in the database")
                await ctx.send(embed = discord.Embed(title="COMMIT SUCCESS", description=desc, color=0x00FF00))
                return  
            
            else:
                desc = (f"Unable to find {user_name} in the database. Use {prefx}roster to verify "
                    "user.")
                await ctx.send(embed = discord.Embed(title="SQL ERROR", description=desc, color=0xFF0000))
                return 

@deletenote.error
async def deletenote_error(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))

@discord_client.command()
async def viewnote(ctx, *, mem):

    # Check server and Member Role
    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    # Get user object 
    member_id = await botAPI.user_converter_db(ctx, mem)

    # If user object can't re resolved then exit 
    if member_id == None:
        desc = (f"Was unable to resolve {mem}. This command supports mentions, "
        "IDs, username and nicknames.")
        await ctx.send(embed = discord.Embed(title="RESOLVE ERROR", description=desc, color=0xFF0000))
        return

    result = dbconn.get_user_byDiscID((member_id,))
    if len(result) == 1:

        # Attempt to get the user member object
        user_obj = await ctx.guild.get_member(member_id)

        # See if you can get the member object
        if isinstance(user_obj, discord.Member):
            user_name = user_obj.display_name
        else:
            user_name = mem

        note = result[0][8]
        ids = []
        search = False
        if re.findall(r"msgID:.\d+", note, re.IGNORECASE):
            for i in re.findall(r"msgID:.\d+", note, re.IGNORECASE):
                ids.append(re.search(r"\d+", i).group())
        if ids:
            search = True
        await ctx.send(f"Current Notes for {user_name}:\n```{note}```")
        if search:
            await ctx.send(f"Message IDs found, would you like those messages retrieved?\n(Yes/No)")
            try:
                response = await discord_client.wait_for('message', check = botAPI.yesno_check, timeout=30)
                if response.content.lower() == 'no':
                    return
            except asyncio.TimeoutError:
                await ctx.send("Just letting you know.. I am no longer waiting for a response")
                return
            
            for msgID in ids:
                zuluServer = discord_client.get_guild(int(config['discord']['zuludisc_id']))
                leaderChannel = zuluServer.get_channel(int(config['discord']['leadernotes']))
                try:
                    await ctx.send("Loading.. ")
                    msg = await leaderChannel.get_message(int(msgID))
                except discord.Forbidden as e:
                    msg = (f"Permission denied to view {leaderChannel.name}\n{e}")
                    await ctx.send(embed = discord.Embed(title="Forbidden Exception", description=msg, color=0xFF0000))
                    return
                except discord.NotFound as e:
                    msg = (f"Message not found")
                    await ctx.send(embed = discord.Embed(title=f"Message not found for id number {msgID}", description=msg, color=0xFF0000))
                    continue
                    
                if msg.attachments:
                    files = []
                    async with aiohttp.ClientSession() as session:
                        for attachment_obj in msg.attachments:
                            async with session.get(attachment_obj.url) as resp:
                                buffer = io.BytesIO(await resp.read())
                                files.append(discord.File(fp=buffer, filename=attachment_obj.filename))
                    files = files or None
                    await ctx.send(f"**Message by:**\n{msg.author.display_name} on {msg.created_at.strftime('%d %b %Y %H:%M').upper()} Zulu\n"
                        f"**Content:**\n{msg.clean_content}", files=files)
                else:
                    await ctx.send(f"**Message by:**\n{msg.author.display_name} on {msg.created_at.strftime('%d %b %Y %H:%M').upper()} Zulu\n"
                        f"**Content:**\n{msg.clean_content}")
            return
    
    else:
        await ctx.send("No results were found, or duplicate results were found. Please checkout logs")

@viewnote.error
async def viewnote_error(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))

@discord_client.command()
async def getmessage(ctx, msgID):
    """ Get messages from leader-notes """

    # Check user and user
    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    if msgID.isdigit() == False:
        desc = (f"Invalid argument {msgID}")
        await ctx.send(embed = discord.Embed(title="RECORD NOT FOUND", description=desc, color=0xFF0000))
        return

    zuluServer = discord_client.get_guild(int(config['discord']['zuludisc_id']))
    leaderChannel = zuluServer.get_channel(int(config['discord']['leadernotes']))
    try:
        await ctx.send("Loading.. ")
        msg = await leaderChannel.get_message(int(msgID))
    except discord.Forbidden as e:
        msg = (f"Permission denied to view {leaderChannel.name}\n{e}")
        await ctx.send(embed = discord.Embed(title="Forbidden Exception", description=msg, color=0xFF0000))
        return
    except discord.NotFound as e:
        msg = (f"Message not found")
        await ctx.send(embed = discord.Embed(title="Message not found", description=msg, color=0xFF0000))
        return
        
    if msg.attachments:
        files = []
        async with aiohttp.ClientSession() as session:
            for attachment_obj in msg.attachments:
                async with session.get(attachment_obj.url) as resp:
                    buffer = io.BytesIO(await resp.read())
                    files.append(discord.File(fp=buffer, filename=attachment_obj.filename))
        files = files or None
        await ctx.send(f"**Message by:**\n{msg.author.display_name} on {msg.created_at.strftime('%d %b %Y %H:%M').upper()} Zulu\n"
            f"**Content:**\n{msg.clean_content}", files=files)
    else:
        await ctx.send(f"**Message by:**\n{msg.author.display_name} on {msg.created_at.strftime('%d %b %Y %H:%M').upper()} Zulu\n"
            f"**Content:**\n{msg.clean_content}")

@getmessage.error
async def getmsg_error(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))



@discord_client.command()
async def caketime(ctx):
    cakePath = "utils/images/cakes"
    ranCake = Path(cakePath).joinpath(random.choice(listdir(cakePath)))
    f = discord.File(str(ranCake), filename=str(ranCake.name))
    if ranCake.suffix == ".mp4":
        await ctx.send(file=f)
    else:
        embed = discord.Embed(title='Nubbz',color=0xFFD700)
        embed.set_footer(text=f"{str(ranCake.name)}")
        embed.set_image(url=f"attachment://{str(ranCake.name)}")
        await ctx.send(embed=embed, file=f)
     

#####################################################################################################################
                                             # Displaying pandas data
#####################################################################################################################
@discord_client.command()
async def export(ctx):

    # Check user and server
    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    # get last sunday integer to calculate last sunday. Then pull from that day and beyond
    today = datetime.utcnow()
    # use 5 minutes as a grace peried for when the db takes too long to update
    lastSunday = (today + timedelta(days=(1 - today.isoweekday()))).replace(hour=0, minute=0, second=0, microsecond=0)
    # Calculate the date range for the sql query
    startDate = (lastSunday - timedelta(weeks=4)).strftime('%Y-%m-%d %H:%M:%S')
    endDate = lastSunday.strftime('%Y-%m-%d %H:%M:%S')
    
    # query
    sql = (f"""
    SELECT 
        MembersTable.Name, 
		Memberstable.Tag, 
		MembersTable.is_Active, 
		DonationsTable.Tag, 
		DonationsTable.increment_date, 
		DonationsTable.Current_Donation 
    FROM 
		MembersTable, DonationsTable 
	WHERE
		MembersTable.Tag = DonationsTable.Tag
	AND
		DonationsTable.increment_date BETWEEN '{startDate}' AND '{endDate}'
	AND
		MembersTable.is_Active = 'True';
       """)
    # Create df out of sql data
    df = pd.read_sql_query(sql, dbconn.conn)
    df['increment_date'] = pd.to_datetime(df['increment_date'], format='%Y-%m-%d %H:%M:%S')

    if df.empty:
        await ctx.send("Not enough data collected to generate a report")
        return
    
    # Remove duplicate Tag column
    #@ts-ignore
    df = df.loc[:,~df.columns.duplicated()]
    
    # Create the date ranges
    mask1 = (df['increment_date'] > (lastSunday - timedelta(days=7))) & (df['increment_date'] < lastSunday)
    mask2 = (df['increment_date'] > (lastSunday - timedelta(days=14))) & (df['increment_date'] < (lastSunday - timedelta(days=7)))
    mask3 = (df['increment_date'] > (lastSunday - timedelta(days=21))) & (df['increment_date'] < (lastSunday - timedelta(days=14)))
    mask4 = (df['increment_date'] > (lastSunday - timedelta(days=28))) & (df['increment_date'] < (lastSunday - timedelta(days=21)))

    # Take the max FIN of each user as a series and convert to our new DF
    df_out = df.loc[mask1].groupby(['Name', 'Tag'])['Current_Donation'].max().reset_index()

    # Set index to the tags instead of the built-int int index
    df_out.set_index('Tag', inplace=True)

    # Change column name of "Current FIN" to the date
    df_out.rename(columns={"Current_Donation":f"{(lastSunday - timedelta(days=1)).strftime('%d%b').upper()}"}, inplace=True)

    # Do the same for the second column
    df_out[f'{(lastSunday - timedelta(days=8)).strftime("%d%b").upper()}'] = df.loc[mask2].groupby(['Name', 'Tag'])['Current_Donation'].max().reset_index().set_index('Tag')['Current_Donation']
    df_out[f'{(lastSunday - timedelta(days=8)).strftime("%d%b").upper()}'] = df_out[f'{(lastSunday - timedelta(days=8)).strftime("%d%b").upper()}'].fillna(0).astype(np.int64)
    # And the third and fourth
    df_out[f'{(lastSunday - timedelta(days=15)).strftime("%d%b").upper()}'] = df.loc[mask3].groupby(['Name', 'Tag'])['Current_Donation'].max().reset_index().set_index('Tag')['Current_Donation']
    df_out[f'{(lastSunday - timedelta(days=15)).strftime("%d%b").upper()}'] = df_out[f'{(lastSunday - timedelta(days=15)).strftime("%d%b").upper()}'].fillna(0).astype(np.int64)
    df_out[f'{(lastSunday - timedelta(days=22)).strftime("%d%b").upper()}'] = df.loc[mask4].groupby(['Name', 'Tag'])['Current_Donation'].max().reset_index().set_index('Tag')['Current_Donation']
    df_out[f'{(lastSunday - timedelta(days=22)).strftime("%d%b").upper()}'] = df_out[f'{(lastSunday - timedelta(days=22)).strftime("%d%b").upper()}'].fillna(0).astype(np.int64)

    # Calculate the difference between the two weeks
    #df_out['Diff'] = df_out.iloc[:,1] - df_out.iloc[:,2]
    df_out['Diff'] = df_out.apply(lambda x: x[1] - x[2] if x[2] > 0 else 0, axis=1)

    # Re order the columns
    cols = df_out.columns.tolist()
    cols.pop(-1)
    cols.insert(1, "Diff")
    df_out = df_out[cols]

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
   
    redFill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='00FF0000'))
    yelFill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='00FFFF00'))
    greFill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='0000FF00'))

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'

    for cell in ws['B']:
        if cell.value == 'Diff':
            continue
        if int(cell.value) < 300:
            if int(ws['D'+str(cell.row)].value) == 0:
                cell.fill = yelFill
            else:
                cell.fill = redFill
        else:
            cell.fill = greFill
    
    wb.save("pandas_openpyxl.xlsx")
    f = discord.File("pandas_openpyxl.xlsx", filename=f'{(lastSunday - timedelta(days=1)).strftime("%d%b").upper()}.xlsx')
    await ctx.send(file=f)

@export.error
async def export_err(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))

@discord_client.command()
async def report(ctx):
    today = datetime.utcnow()
    #lastSunday = (today + timedelta(days=(1 - today.isoweekday()))).replace(hour=1, minute=0, second=0, microsecond=0)
    lastSunday = (today + timedelta(days=(1 - today.isoweekday()))).replace(hour=0, minute=0, second=0, microsecond=0)
    today = today.strftime('%Y-%m-%d %H:%M:%S')
    startDate = (lastSunday - timedelta(weeks=1)).strftime('%Y-%m-%d %H:%M:%S')

    sql = (f"""
        SELECT MembersTable.Name, 
            Memberstable.Tag, 
            MembersTable.is_Active, 
            DonationsTable.Tag, 
            DonationsTable.increment_date, 
            DonationsTable.Current_Donation 
        FROM 
            MembersTable, DonationsTable 
        WHERE
            MembersTable.Tag = DonationsTable.Tag
        AND
            DonationsTable.increment_date BETWEEN '{startDate}' AND '{today}'
        AND
            MembersTable.is_Active = 'True';
        """)

    # read SQL then convert date to tdate
    df = pd.read_sql_query(sql, dbconn.conn)
    df['increment_date'] = pd.to_datetime(df['increment_date'], format='%Y-%m-%d %H:%M:%S')

    # Remove duplicate Tag column
    df = df.loc[:,~df.columns.duplicated()]

    # First make the two masks
    before_sun = df['increment_date'] <= lastSunday
    after_sun = df['increment_date'] >= lastSunday   

    # Calculate the diff for this week and save it to its own DF
    # Rename column, reset index
    df_out = df.loc[after_sun].groupby(['Tag', 'Name'])['Current_Donation'].agg(['min','max']).diff(axis=1)
    # Exit if there isn't enough data
    if df_out.empty:
        await ctx.send("Please wait an hour to accurately calculate your donations. Thank you for your patience.")
        return
    df_out.drop('min', axis=1, inplace=True)
    df_out.rename(columns={'max':'Current'}, inplace=True)
    df_out.reset_index(inplace=True)
    df_out.set_index('Tag', inplace=True)

    # Create current FIN column
    df_out['Current_FIN'] = df.loc[after_sun].groupby(['Tag'])['Current_Donation'].max()

    # create last sunday column
    df_out[f'{(lastSunday - timedelta(days=1)).strftime("%d%b").upper()}'] = df.loc[before_sun].groupby(['Tag'])['Current_Donation'].max()

    # Clean up data change NaN and Float to 
    df_out[df_out.columns[1:]] = df_out[df_out.columns[1:]].fillna(0).astype(np.int64)

    # Sort names column
    #df_out.sort_values(by='Name.Upper', inplace=True)
    df_out = df_out.iloc[df_out.Name.str.lower().argsort()]

    # Dataframe to html
    html = df_out.to_html(index=False, justify="center")
    # load into a BS object
    soup = bs4.BeautifulSoup(html, "lxml")
    # extract the table
    table = soup.find("table")

    scriptTag = """// Query for the table tags and coloring to the table
const tableElm = document.getElementsByTagName("table")[0]; 
for (const row of tableElm.rows) {
  const childToStyle = row.children[1];
  console.log(childToStyle.textContent);
  if (Number(childToStyle.textContent) < 300) { 
    childToStyle.classList.add("redClass");
  } else if (Number(childToStyle.textContent) > 299) {
      childToStyle.classList.add("greenClass")
  }
}  """

    cssTag = """.redClass {
                    background-color: red;
                    font-weight : bold;
                    }
                    .greenClass {
                        background-color : green;
                        font-weight : bold;
                    }"""

    base = (f"""<!DOCTYPE html>
                <html>
                <head>
                    <meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=0"> 
                    <title>Reddit Zulu</title>
                    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/normalize.css@8.0.1/normalize.css"/>
                    <style>
                    {cssTag}
                    </style>
                </head>
                <body>
                {table}
                <script type="text/javascript">
                {scriptTag}
                </script>
                </body>
                </html>""") #.format(cssTag, table, scriptTag)

    soup = bs4.BeautifulSoup(base, "lxml")
    #new_link = soup.new_tag("script", src="pandamod.js")
    #soup.html.append(new_link)

    with open("utils/web/report.html", "w", encoding="utf-8") as outfile:
        outfile.write(str(soup))

    f = discord.File("utils/web/report.html", filename="report.html")
    await ctx.send(file=f)
    await ctx.send(f"Keep in mind that the database only updates every 15 minutes.")
    return

@discord_client.command(aliases=["q"])
async def queue(ctx):
    apps = dbconn.get_apps()
    app_u = []
    app_m = ""
    for app in apps:
        if app[0] not in app_u:
            app_u.append(app[0])
            app_m += f"{app[0]:<15} {app[1]}\n"
    await ctx.send(app_m)

@discord_client.command(aliases=["t", "T"])
async def top(ctx, arg=None):
    # Set dates
    start_date = botAPI.last_sunday().strftime("%Y-%m-%d %H:%M:%S")
    end_date = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

    # Initialize list
    top_stats = []

    # Update all data first
    async with ctx.typing():
        await udt.update_donationstable(dbconn, coc_client2)

    # gather data
    users = dbconn.get_all_active()
    for user in users:
        users_week = dbconn.get_Donations((user[0], start_date, end_date))
        #print(users_week[0], users_week[-1])
        top_stats.append(user_tops.Tops(
            user[1],
            user[0],
            user[2],
            users_week[0][4],
            users_week[-1][4],
            users_week[0][2],
            users_week[-1][2]
        ))
    
    if arg == None or arg.lower() in ["-t", "--trophes"]:
        # sort list by top trophies
        top_stats.sort(key = lambda x: x.e_trophy, reverse=True)

        # Set up output
        output = "**Top Trophies:**\n"
        output += (f"`{'rk':<2} {'th':<2} {'trop':<4} {'diff':>5}`\n")
        count = 1
        for user in top_stats:
            output += (
                f"`{count:<2} {user.townhall:<2} {user.e_trophy:<4} {user.e_trophy - user.s_trophy:>5} {user.name}`\n"
            )
            count +=1
        await ctx.send(output)

    elif arg.lower() in ["-d", "--donation"]:
        # sort list by top donation
        top_stats.sort(key = lambda x: x.e_donation, reverse=True)

        # Set up output
        output = "**Top Donations:**\n"
        output += (f"`{'th':<2} {'don':<4} {'diff':>5}`\n")
        count = 1
        for user in top_stats:
            if count == 10:
                break
            output += (
                f"`{user.townhall:<2} {user.e_donation:<4} {user.e_donation - user.s_donation:>5} {user.name}`\n"
            )
            count +=1
            
        await ctx.send(output)

@discord_client.command()
async def test(ctx):
    dg = discord_client.get_guild(int(config["discord"]["zuludisc_id"]))
    print(dg)
    ud = dg.get_member(328247925786279940)
    print(ud)


#####################################################################################################################
                                             # Loops & Kill Command
#####################################################################################################################
@discord_client.command()
async def killbot(ctx):
    """ Send kill signal to bot to properly close down databse and config file """

    if await botAPI.rightServer(ctx, config) and await botAPI.authorized(ctx, config):
        pass
    else:
        return

    await ctx.send("Tearing down, please hold.")
    await ctx.send("Closing database..")
    dbconn.conn.close()
    await ctx.send("Terminating bot..")
    await ctx.send("_Later._")
    await discord_client.logout()

@killbot.error
async def killbot_error(ctx, error):
    await ctx.send(embed = discord.Embed(title="ERROR", description=error.__str__(), color=0xFF0000))

@discord_client.event
async def on_message(message):
    if message.channel.id == 293953660059385857: # replace with leaders chat
        if message.content.startswith("Application"):
            data = message.content.split(":")[1]
            user, tag = data.split("(")
            user = user.strip()
            tag = tag.rstrip(")")
            dbconn.new_app((
                tag,
                user,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ))
    await discord_client.process_commands(message)


async def weeklyRefresh(discord_client, botMode):
    """ Function used to update the databsae with new data """

    # Don't allow the bot to loop when in devMode
    if botMode == "devBot":
        print("Running in dev mode, disabling database update.")
        return
    await discord_client.wait_until_ready()
    while not discord_client.is_closed():
        # Calculate the wait time in minute for next "top of hour"
        wait_time = 60 - datetime.utcnow().minute
        if wait_time <= 15:
            pass
        elif wait_time <= 30:
            wait_time = wait_time - 15
        elif wait_time <= 45:
            wait_time = wait_time - 30
        else:
            wait_time = wait_time - 45

        print(f"\n\nWaiting {wait_time} minutes until next update.")
        await asyncio.sleep(wait_time * 60) #60

        # Update message every time we update db
        game = Game("Updating Donations")
        await discord_client.change_presence(status=discord.Status.dnd, activity=game)

        guild = discord_client.get_guild(int(config[botMode]['guild_lock']))
        # Get all users in the database
        get_all = dbconn.get_all_active()

        # Update all donations
        await udt.update_donationstable(dbconn, coc_client2)

        # See if the users are still part of the clan
        user = ''
        for user in get_all:
            # if mem in planning server
            in_planning = ""
            if int(user[4]) in (mem.id for mem in discord_client.get_guild(int(config['discord']['plandisc_id'])).members):
                if user[6] == "True":
                    pass
                else:
                    in_planning = "True"
            else:
                if user[6] == "False":
                    pass
                else:
                    in_planning = "False"

            # Grab the users CoC stats to see if there is any updates needed on their row
            try:
                res = coc_client.get_member(user[0])
            except:
                print(f"Could not retrive clash member {user[0]} data")
                await (discord_client.get_channel(int(config["discord"]["thelawn"]))).send(f"Could not retrive clash member {user[0]} data")
                continue

            if isinstance(res, Response) == False:
                print(f"Could not retrive clash member {user[0]} {user[1]} data. Returned a None object")
                await (discord_client.get_channel(int(config["discord"]["thelawn"]))).send(f"Could not retrive clash member {user[0]} data. Returned a None object")
                continue

            if res.status_code != 200:
                  print(f"Could not connect to CoC API with {user[0]}")
                  await (discord_client.get_channel(int(config["discord"]["thelawn"]))).send(f"Could not connect to CoC API with {user[0]}")
                  continue

            # Instantiate the users clash data
            try:
                mem_stats = clash_stats.ClashStats(res.json())
            except:
                print(f"Could not instantiate ClashStat object: {user[0]} {user[1]}")
                await (discord_client.get_channel(int(config["discord"]["thelawn"]))).send(f"Could not instantiate ClashStat object: {user[0]} {user[1]}")
                continue

            # Grab the users discord object and the object for the TH role
            exists, disc_UserObj = botAPI.is_DiscordUser(guild, config, user[4])

            if exists == False:
                print(f"User does not exist {user[1]} does not exist in this server")
                await (discord_client.get_channel(int(config["discord"]["thelawn"]))).send(f"User does not exist {user[1]} does not exist in this server")
                continue

            # Grab users role object
            current_roleObj = botAPI.get_townhallRole(guild, mem_stats.townHallLevel)

            # find if their TH role has changed
            assigned_roleObjs =[ role for role in disc_UserObj.roles if role.name.startswith('th') ]

            # Check if users th level object is already assigned to them
            if current_roleObj in assigned_roleObjs:
                # they have the role, make sure they don't have more than one
                if len(assigned_roleObjs) == 1:
                    pass
                else:
                    await disc_UserObj.remove_roles(*assigned_roleObjs)
                    await disc_UserObj.add_roles(current_roleObj)

            # if they don't have the role check if they have any th roles
            elif len(assigned_roleObjs) > 0:
                await disc_UserObj.remove_roles(*assigned_roleObjs)
                await disc_UserObj.add_roles(current_roleObj)
            # finally, just add the role if they don't have any
            else:
                await disc_UserObj.add_roles(current_roleObj)

            # Check to see if they are current in zulu or somewhere else
            in_zulu = "False"
            if mem_stats.clan_name == "Reddit Zulu":
                in_zulu = "True"
            else:
                in_zulu = "False"
            dbconn.update_donations((
                    datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
                    mem_stats.tag,
                    mem_stats.achieve["Friend in Need"]['value'],
                    in_zulu,
                    mem_stats.trophies
                ))

            # update users table  in_planning
            #(TownHallLevel, League, inPlanningServer, Tag)
            dbconn.update_members_table((mem_stats.townHallLevel, mem_stats.league_name, in_planning, mem_stats.tag))

        # reset message
        messages = [
            (discord.ActivityType.listening ,   "Spotify"),
            (discord.ActivityType.playing   ,   "Overwatch"),
            (discord.ActivityType.playing   ,   "Clash of Clans"),
            (discord.ActivityType.playing   ,   "with cat nip~"),
            (discord.ActivityType.watching ,   "Fairy Tail"),
            (discord.ActivityType.playing   ,   "I'm not a cat!"),
            (discord.ActivityType.watching  ,   "panther.help")
        ]

        activ = random.choice(messages)
        activity = discord.Activity(type = activ[0], name=activ[1])
        await discord_client.change_presence(status=discord.Status.online, activity=activity)


if __name__ == "__main__":
    #discord_client.loop.create_task(weeklyRefresh(discord_client, botMode))
    discord_client.run(config[botMode]['bot_token'])
    # print(discord_client.guilds)
    # print('wttfff')
    # update = d_sync.UpdateLoop(discord_client, dbconn, botMode, coc_client2, config) 
    # discord_client.loop.create_task(update.run())
